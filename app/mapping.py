"""
Lecture de l'onglet ``mapping`` du fichier Excel et construction de la
**séquence de diapositives** à générer (titres de section insérés
automatiquement là où un numéro est absent).

Types de slide / index dans ``template_slide.pptx`` (ordre imposé par le
gabarit, voir README) :

- ``TITLE_SECTION``   : diapo titre de section (pas de graphique)
- ``GRAPH_VERTI``     : 1 graphique, gabarit « verticale »
- ``GRAPH_HORI``      : 1 graphique, gabarit « horizontale »
- ``GRAPH_2`` / ``GRAPH_3`` / ``GRAPH_4`` : 2, 3 ou 4 graphiques sur la
  même diapo (même nombre de placeholders chart sur le layout)

Règles :

1. ``no_slide`` = numéro (1..N) de la diapo cible dans le livrable final.
2. Plusieurs lignes du mapping peuvent partager le même ``no_slide`` :
   chaque ligne = 1 graphique sur la slide. **Ou** une seule ligne avec
   la colonne ``split_charts_by`` (ex. ``segment``) : le moteur duplique
   automatiquement la ligne en un graphique par valeur unique lue dans
   l'onglet ``slide_N`` (filtre imposé par ``axis_layout`` — voir README).
3. Si, une fois les lignes ordonnées par ``no_slide``, il manque un
   numéro entre 1 et ``max(no_slide)``, une diapo ``TITLE_SECTION`` est
   insérée automatiquement à cette position.
4. Gabarit pour une diapo avec graphique(s) :

   - **1 ligne** : ``vertical`` dans ``chart_type`` de la 1re ligne →
     ``GRAPH_VERTI`` ; sinon ``GRAPH_HORI``
   - **2, 3 ou 4 lignes** (même ``no_slide``) → ``slide_2_graphs``,
     ``slide_3_graphs`` ou ``slide_4_graphs`` (autant de placeholders
     chart sur la diapo que de lignes). Au-delà de 4 : erreur explicite.
"""

from __future__ import annotations

from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# Entier = index 0-based dans la banque template_slide.pptx (ordre des
# 6 diapos témoin : title, graph verti, graph hori, 2, 3, 4 graphiques).
TITLE_SECTION: int = 0
GRAPH_VERTI: int = 1
GRAPH_HORI: int = 2
GRAPH_2: int = 3
GRAPH_3: int = 4
GRAPH_4: int = 5

SLIDE_TYPE_NAMES: dict[int, str] = {
    TITLE_SECTION: "title_section",
    GRAPH_VERTI: "graph_verti",
    GRAPH_HORI: "graph_hori",
    GRAPH_2: "slide_2_graphs",
    GRAPH_3: "slide_3_graphs",
    GRAPH_4: "slide_4_graphs",
}
MAX_GRAPHS_PER_SLIDE: int = 4


@dataclass(frozen=True)
class MappingRow:
    """Une ligne de l'onglet ``mapping`` = un graphique à insérer."""

    no_slide: int
    column_pct: str
    chart_type: str
    axis_layout: str
    filter_brand: str
    filter_modalite: str
    filter_segment: str
    sort_order: str
    sort_by: str
    titre_graphique: str
    series_order: str
    # Si non vide : une seule ligne pour ce no_slide; expansion auto en N
    # graphiques, un par valeur de la dimension (dérivée de slide_N).
    split_charts_by: str = ""
    # Si renseigné (> 0) : ne garder que les N premières catégories après tri.
    top_n: int | None = None


@dataclass(frozen=True)
class SlideSpec:
    """Une diapositive à générer, dans l'ordre final du deck."""

    position: int                 # 1..N : position dans le livrable
    slide_type: int               # TITLE_SECTION / GRAPH_VERTI / GRAPH_HORI
    rows: tuple[MappingRow, ...]  # graphiques à placer (vide si TITLE_SECTION)


def _norm(s: Any) -> str:
    return str(s).strip().lower().replace(" ", "_")


def parse_mapping(path: str | Path) -> list[MappingRow]:
    """Lit l'onglet « mapping » et renvoie une liste de ``MappingRow``."""
    wb = load_workbook(path, read_only=True, data_only=True)
    if "mapping" not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Onglet « mapping » absent. Onglets trouvés : {wb.sheetnames}"
        )
    ws = wb["mapping"]

    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h is not None:
            headers[_norm(h)] = c

    def _col(*aliases: str) -> int | None:
        for a in aliases:
            k = _norm(a)
            if k in headers:
                return headers[k]
        return None

    c_slide  = _col("no_slide", "slide", "slide_id")
    c_pct    = _col("column_pct", "colonne_pct", "pct_column")
    c_chart  = _col("chart_type", "chart")
    c_title  = _col("titre_graphique", "chart_title", "titre")
    c_sorder = _col("series_order", "ordre_series")
    c_layout = _col("axis_layout", "layout", "pivot")
    c_fbrand = _col("filter_brand", "filtre_marque", "marque")
    c_fmod   = _col("filter_modalite", "filter_modality", "filtre_modalite", "modalite")
    c_fseg   = _col("filter_segment", "filtre_segment", "segment")
    c_sord   = _col("sort_order", "sort_direction", "tri", "ordre_tri")
    c_sby    = _col("sort_by", "tri_par", "sort_series")
    c_split  = _col("split_charts_by", "split_charts", "repartir_par", "un_graphique_par", "one_chart_per")
    c_topn   = _col("top_n", "top", "n_first", "limit_categories", "nb_categories")

    if not c_slide:
        wb.close()
        raise ValueError("Colonne « no_slide » introuvable dans l'onglet mapping.")

    def _cell(c: int | None, row_idx: int, default: str = "") -> str:
        if not c:
            return default
        x = ws.cell(row=row_idx, column=c).value
        return default if x is None or str(x).strip() == "" else str(x).strip()

    rows: list[MappingRow] = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=c_slide).value
        if v is None or str(v).strip() == "":
            continue

        pct     = _cell(c_pct, r)
        ct      = _cell(c_chart, r, "bar_clustered_vertical")
        layout  = _cell(c_layout, r)
        f_brand = _cell(c_fbrand, r)
        f_mod   = _cell(c_fmod, r)
        f_seg   = _cell(c_fseg, r)
        title   = _cell(c_title, r)
        sorder  = _cell(c_sorder, r)
        s_ord   = _cell(c_sord, r).lower()
        s_by    = _cell(c_sby, r)
        split_by = _cell(c_split, r).lower() if c_split else ""

        top_n_val: int | None = None
        if c_topn:
            tx = ws.cell(row=r, column=c_topn).value
            if tx is not None and str(tx).strip() != "":
                try:
                    tn = int(float(tx))
                except (TypeError, ValueError) as e:
                    wb.close()
                    raise ValueError(
                        f"Ligne mapping r={r} : « top_n » invalide ({tx!r}). "
                        "Attendu : un entier positif ou vide."
                    ) from e
                if tn < 0:
                    wb.close()
                    raise ValueError(f"Ligne mapping r={r} : top_n doit être ≥ 0 (reçu {tn}).")
                top_n_val = tn if tn > 0 else None

        if not layout:
            wb.close()
            raise ValueError(
                f"Slide {v} : colonne « axis_layout » manquante ou vide. "
                "Valeurs attendues : brands_x_modalites, brands_x_segments, "
                "modalites_x_brands, modalites_x_segments, "
                "segments_x_brands, segments_x_modalites."
            )

        if s_ord and s_ord not in ("asc", "desc", "none"):
            wb.close()
            raise ValueError(
                f"Slide {v} : sort_order = « {s_ord} » invalide. "
                "Valeurs autorisées : asc, desc, none (ou vide pour auto)."
            )

        rows.append(MappingRow(
            no_slide=int(v),
            column_pct=pct,
            chart_type=ct,
            axis_layout=layout,
            filter_brand=f_brand,
            filter_modalite=f_mod,
            filter_segment=f_seg,
            sort_order=s_ord,
            sort_by=s_by,
            titre_graphique=title,
            series_order=sorder,
            split_charts_by=split_by,
            top_n=top_n_val,
        ))

    wb.close()
    return rows


def _open_filter(f: str) -> bool:
    return not f or not str(f).strip() or str(f).strip().lower() in ("all", "")


def _split_string_to_dimension(s: str) -> str:
    """``segment`` | ``brand`` | ``modality`` (synonymes FR/EN)."""
    if not s or not str(s).strip():
        return ""
    k = _norm(s)
    if k in ("segment", "segments", "seg", "serie", "series"):
        return "segment"
    if k in ("brand", "brands", "marque", "marques"):
        return "brand"
    if k in (
        "modality",
        "modalities",
        "modalite",
        "modalites",
        "sous_categorie",
        "sous_categories",
    ):
        return "modality"
    raise ValueError(
        f"split_charts_by = {s!r} : utiliser segment, brand ou modality "
        f"(les synonymes marque / modalité / segment sont reconnus)."
    )


def expand_split_charts(excel_path: str | Path, rows: list[MappingRow]) -> list[MappingRow]:
    """
    Double chaque ligne dont ``split_charts_by`` est renseigné : une entrée
    par valeur **unique** de la dimension cible, lue dans l'onglet
    ``slide_{no_slide}`` (même règles de filtre que le pivot : la dimension
    doit être celle imposée par ``axis_layout``).

    - La ligne source doit être **seule** pour son ``no_slide``.
    - Le filtre sur cette dimension (``filter_brand`` / ``filter_modalite`` /
      ``filter_segment``) doit être vide ou ``all``.
    - Nombre de graphiques plafonné par :data:`MAX_GRAPHS_PER_SLIDE` (4).

    Dans ``titre_graphique``, le texte ``{value}`` est remplacé par la valeur
    courante (ex. le nom du segment).
    """
    from .slide_data import (
        get_other_axis_dimension,
        list_unique_values_for_dimension,
        parse_slide_sheet,
    )

    excel_path = Path(excel_path)
    groups: dict[int, list[MappingRow]] = {}
    for r in rows:
        groups.setdefault(r.no_slide, []).append(r)

    out: list[MappingRow] = []
    for no in sorted(groups.keys()):
        grp = groups[no]
        with_split = [r for r in grp if (r.split_charts_by or "").strip()]
        if not with_split:
            out.extend(grp)
            continue
        if len(grp) != 1 or len(with_split) != 1:
            raise ValueError(
                f"no_slide={no} : avec « split_charts_by », il faut exactement **une** ligne "
                f"pour ce numéro (reçu {len(grp)}). Supprimez les doublons ou desactivez split."
            )
        row = with_split[0]
        dim = _split_string_to_dimension(row.split_charts_by)
        other = get_other_axis_dimension(row.axis_layout)
        if dim != other:
            raise ValueError(
                f"no_slide={no} : pour axis_layout={row.axis_layout!r}, repartir les graphiques "
                f"selon la dimension **{other}** (celle a filtrer sur le cube). "
                f"Indiquez split_charts_by = {other!r} (ou marque / modalite / segment selon le cas)."
            )
        if dim == "brand" and not _open_filter(row.filter_brand):
            raise ValueError(
                f"no_slide={no} : laissez filter_brand vide (ou all) quand split_charts_by=brand."
            )
        if dim == "modality" and not _open_filter(row.filter_modalite):
            raise ValueError(
                f"no_slide={no} : laissez filter_modalite vide (ou all) quand split_charts_by=modality."
            )
        if dim == "segment" and not _open_filter(row.filter_segment):
            raise ValueError(
                f"no_slide={no} : laissez filter_segment vide (ou all) quand split_charts_by=segment."
            )

        data = parse_slide_sheet(excel_path, row.no_slide, column_pct=row.column_pct)
        values = list_unique_values_for_dimension(data, dim)
        if not values:
            raise ValueError(
                f"no_slide={no} : aucune valeur pour la dimension {dim!r} dans l'onglet slide_{row.no_slide}."
            )
        if len(values) > MAX_GRAPHS_PER_SLIDE:
            raise ValueError(
                f"no_slide={no} : {len(values)} graphique(s) a generer, maximum "
                f"{MAX_GRAPHS_PER_SLIDE} sur une diapo. Reduisez les donnees ou utilisez des slides separees."
            )
        for v in values:
            if dim == "brand":
                fb, fm, fs = str(v), row.filter_modalite, row.filter_segment
            elif dim == "modality":
                fb, fm, fs = row.filter_brand, str(v), row.filter_segment
            else:
                fb, fm, fs = row.filter_brand, row.filter_modalite, str(v)
            title = (row.titre_graphique or "").replace("{value}", str(v))
            out.append(
                replace(
                    row,
                    filter_brand=fb,
                    filter_modalite=fm,
                    filter_segment=fs,
                    titre_graphique=title,
                    split_charts_by="",
                )
            )
        print(
            f"[INFO] no_slide={no} : split_charts_by={dim!r} -> {len(values)} graphique(s) : {values!r}"
        )
    return out


def detect_slide_type(rows: list[MappingRow]) -> int:
    """
    Choisit l'index de la diapo modèle dans ``template_slide.pptx``.

    - 0 graphique  → :data:`TITLE_SECTION`
    - 1 graphique  → verti ou hori selon ``chart_type`` de la 1re ligne
    - 2 / 3 / 4    → :data:`GRAPH_2` / :data:`GRAPH_3` / :data:`GRAPH_4`

    Lève :class:`ValueError` si plus de :data:`MAX_GRAPHS_PER_SLIDE` lignes
    partagent le même ``no_slide``.
    """
    if not rows:
        return TITLE_SECTION
    n = len(rows)
    if n == 1:
        first = (rows[0].chart_type or "").strip().lower()
        return GRAPH_VERTI if "vertical" in first else GRAPH_HORI
    if n == 2:
        return GRAPH_2
    if n == 3:
        return GRAPH_3
    if n == 4:
        return GRAPH_4
    raise ValueError(
        f"Jusqu'a {MAX_GRAPHS_PER_SLIDE} graphique(s) par no_slide. "
        f"Reçu {n} ligne(s) — réduire le nombre de lignes pour ce numéro "
        f"ou étendre le gabarit."
    )


def build_slide_sequence(rows: list[MappingRow]) -> list[SlideSpec]:
    """
    Construit la séquence ordonnée des diapositives à générer.

    - Regroupe les ``MappingRow`` par ``no_slide`` (préserve l'ordre du
      fichier Excel pour les lignes du même numéro).
    - Parcourt 1..N (N = ``max(no_slide)``) ; à chaque position :

      - s'il existe des lignes → gabarit selon le nombre de lignes / verti-hori
      - sinon → ``TITLE_SECTION`` (diapo auto-insérée)
    """
    if not rows:
        return []

    groups: dict[int, list[MappingRow]] = {}
    for row in rows:
        groups.setdefault(row.no_slide, []).append(row)

    max_no = max(groups.keys())
    sequence: list[SlideSpec] = []
    for pos in range(1, max_no + 1):
        grp = groups.get(pos)
        if grp:
            sequence.append(
                SlideSpec(
                    position=pos,
                    slide_type=detect_slide_type(grp),
                    rows=tuple(grp),
                )
            )
        else:
            sequence.append(
                SlideSpec(position=pos, slide_type=TITLE_SECTION, rows=())
            )
    return sequence
