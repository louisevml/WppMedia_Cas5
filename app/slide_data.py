"""
Lecture des onglets ``slide_{N}`` (données) et transformations génériques :

- :class:`SlideData`        : représentation brute d'un onglet
- :func:`parse_slide_sheet` : lit l'onglet Excel ``slide_N``
- :func:`apply_axis_layout` : pivot (catégories × séries) + filtres
- :func:`sort_categories`   : tri des catégories selon une série
- :func:`truncate_categories_top_n` : garde les N premières catégories (mapping ``top_n``)
- :func:`reorder_series`    : réordonne / filtre les séries
- :func:`build_chart_data`  : prépare le ``CategoryChartData`` prêt à insérer
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pptx.chart.data import CategoryChartData


@dataclass
class SlideData:
    """Résultat du parsing d'un onglet ``slide_N``."""

    series_names: list[str]
    categories: list[str]
    values_matrix: list[list[float]]   # values_matrix[serie_idx][cat_idx]
    brands: list[str] | None = None
    subcategories: list[str] | None = None


# ═════════════════════════════════════════════════════════════════════════════
# 1. parse_slide_sheet  —  en-têtes 3 niveaux + données
# ═════════════════════════════════════════════════════════════════════════════

def _pct_col_indices(column_pct_label: str) -> tuple[int, ...]:
    """
    Retourne les indices 1-based des colonnes % à lire pour chaque bloc
    (structure : 4 blocs × 3 colonnes — Effectif, Nb.colonnes %, Nb.total %).
    """
    key = column_pct_label.strip().lower().replace(" ", "")
    if "total" in key:
        return (5, 8, 11, 14)
    return (4, 7, 10, 13)


def _read_series_labels(ws: Worksheet) -> list[str]:
    """Extrait les noms des séries depuis les lignes 1-2 des en-têtes."""
    r1c = ws.cell(row=1, column=3).value
    r2f = ws.cell(row=2, column=6).value
    r2i = ws.cell(row=2, column=9).value
    r2l = ws.cell(row=2, column=12).value

    def _s(v: Any, fallback: str) -> str:
        if v is None or str(v).strip() == "":
            return fallback
        return str(v).strip()[:40]

    return [
        _s(r1c, "Ensemble"),
        _s(r2f, "Segment 2"),
        _s(r2i, "Segment 3"),
        _s(r2l, "Segment 4"),
    ]


def parse_slide_sheet(
    path: str | Path,
    no_slide: int,
    *,
    column_pct: str,
) -> SlideData:
    """Parse l'onglet ``slide_{no_slide}`` et renvoie un :class:`SlideData`."""
    wb = load_workbook(path, read_only=False, data_only=True)
    sheet_name = f"slide_{no_slide}"
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Onglet « {sheet_name} » absent. Onglets : {wb.sheetnames}")
    ws = wb[sheet_name]

    series_names = _read_series_labels(ws)
    pct_cols = _pct_col_indices(column_pct)
    n_series = len(pct_cols)

    brands: list[str] = []
    subcategories: list[str] = []
    matrix: list[list[float]] = [[] for _ in range(n_series)]

    current_brand = ""
    for r in range(4, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if a is None and b is None:
            continue

        if a is not None and str(a).strip():
            current_brand = str(a).strip()
        subcat = str(b).strip() if b is not None and str(b).strip() else ""

        if current_brand.lower() == "total":
            continue
        if subcat.lower() == "total":
            continue

        brands.append(current_brand)
        subcategories.append(subcat)

        for si, col_idx in enumerate(pct_cols):
            raw = ws.cell(row=r, column=col_idx).value
            if raw is None:
                matrix[si].append(0.0)
            else:
                v = float(raw)
                matrix[si].append(v * 100.0 if abs(v) <= 1.000001 else v)

    wb.close()

    categories = [sub if sub else brand for brand, sub in zip(brands, subcategories)]

    return SlideData(
        series_names=series_names,
        categories=categories,
        values_matrix=matrix,
        brands=brands,
        subcategories=subcategories,
    )


# ═════════════════════════════════════════════════════════════════════════════
# 2. sort_categories — tri des catégories par la valeur d'une série
# ═════════════════════════════════════════════════════════════════════════════

def _resolve_sort_series_index(series_names: list[str], sort_by: str) -> int:
    """Retourne l'index (0-based) de la série désignée par ``sort_by``."""
    if not sort_by or not sort_by.strip():
        return 0

    s = sort_by.strip()
    try:
        idx = int(s)
        if 0 <= idx < len(series_names):
            return idx
        raise ValueError(
            f"sort_by = {idx} hors plage. "
            f"Séries disponibles (0..{len(series_names) - 1}) : {series_names}"
        )
    except ValueError as e:
        if "hors plage" in str(e):
            raise

    low = s.lower()
    normalized = [sn.strip().lower() for sn in series_names]
    for i, n in enumerate(normalized):
        if n == low:
            return i
    for i, n in enumerate(normalized):
        if n.startswith(low):
            return i
    for i, n in enumerate(normalized):
        if low in n:
            return i

    raise ValueError(
        f"sort_by = « {sort_by} » ne correspond à aucune série. "
        f"Séries disponibles : {series_names}"
    )


def sort_categories(
    data: SlideData,
    *,
    sort_order: str,
    sort_by: str,
    chart_type: str,
) -> SlideData:
    """Tri des catégories selon une série donnée."""
    direction = (sort_order or "").strip().lower()
    if direction == "none":
        return data
    ct = chart_type.strip().lower()
    if direction == "":
        # Secteurs : sans tri explicite, conserver l'ordre des modalités
        # issu du tableur (légende partagée, échelle d'opinion cohérente).
        if ct in ("pie", "doughnut"):
            return data
        direction = "asc" if "horizontal" in ct else "desc"
    if direction not in ("asc", "desc"):
        raise ValueError(
            f"sort_order invalide : « {sort_order} ». Attendu : asc, desc, none."
        )

    if not data.categories:
        return data

    series_idx = _resolve_sort_series_index(data.series_names, sort_by)

    order = sorted(
        range(len(data.categories)),
        key=lambda i: data.values_matrix[series_idx][i],
        reverse=(direction == "desc"),
    )
    return SlideData(
        series_names=data.series_names,
        categories=[data.categories[i] for i in order],
        values_matrix=[[row[i] for i in order] for row in data.values_matrix],
        brands=[data.brands[i] for i in order] if data.brands else None,
        subcategories=[data.subcategories[i] for i in order] if data.subcategories else None,
    )


def truncate_categories_top_n(data: SlideData, top_n: int | None) -> SlideData:
    """
    Ne conserve que les ``top_n`` premières catégories (après tri / pivot).

    - ``top_n`` vide, ``None`` ou ``<= 0`` : aucune coupe (tout est gardé).
    - Sinon : tronque les listes parallèles aux ``top_n`` premiers indices.
    """
    if top_n is None or top_n <= 0:
        return data
    if len(data.categories) <= top_n:
        return data
    indices = list(range(top_n))
    return SlideData(
        series_names=data.series_names,
        categories=[data.categories[i] for i in indices],
        values_matrix=[[row[i] for i in indices] for row in data.values_matrix],
        brands=[data.brands[i] for i in indices] if data.brands else None,
        subcategories=[data.subcategories[i] for i in indices] if data.subcategories else None,
    )


# ═════════════════════════════════════════════════════════════════════════════
# 3. reorder_series — réordonne / filtre les séries selon series_order
# ═════════════════════════════════════════════════════════════════════════════

def reorder_series(data: SlideData, series_order: str) -> SlideData:
    """Réordonne (et filtre) les séries selon une liste de noms séparés par virgules."""
    if not series_order or not series_order.strip():
        return data

    requested = [s.strip() for s in series_order.split(",") if s.strip()]
    if not requested:
        return data

    name_to_idx: dict[str, int] = {}
    for i, sn in enumerate(data.series_names):
        name_to_idx[sn.strip().lower()] = i

    new_names: list[str] = []
    new_matrix: list[list[float]] = []
    for req in requested:
        idx = name_to_idx.get(req.strip().lower())
        if idx is not None:
            new_names.append(data.series_names[idx])
            new_matrix.append(data.values_matrix[idx])
        else:
            print(f"[WARN] series_order : « {req} » introuvable, ignoré.")

    if not new_names:
        return data

    return SlideData(
        series_names=new_names,
        categories=data.categories,
        values_matrix=new_matrix,
        brands=data.brands,
        subcategories=data.subcategories,
    )


# ═════════════════════════════════════════════════════════════════════════════
# 4. apply_axis_layout — pivot générique à partir d'un cube 3D
# ═════════════════════════════════════════════════════════════════════════════

_DIM_ALIASES: dict[str, str] = {
    "brand": "brand", "brands": "brand",
    "marque": "brand", "marques": "brand",
    "modalite": "modality", "modalites": "modality",
    "modality": "modality", "modalities": "modality",
    "sous_categorie": "modality", "sous_categories": "modality",
    "segment": "segment", "segments": "segment",
    "cible": "segment", "cibles": "segment",
    "groupe": "segment", "groupes": "segment",
}


def _normalize_dim(s: str) -> str:
    key = s.strip().lower().replace(" ", "_")
    if key not in _DIM_ALIASES:
        raise ValueError(
            f"Dimension inconnue : « {s} ». "
            f"Valeurs attendues : brands / modalites / segments."
        )
    return _DIM_ALIASES[key]


def get_other_axis_dimension(axis_layout: str) -> str:
    """
    Troisième dimension du cube (``brand`` | ``modality`` | ``segment``) pour
    l'``axis_layout`` donné : celle à filtrer à une valeur par graphique quand
    les deux autres dimensions sont en abscisse / légende.
    """
    parts = axis_layout.strip().lower().split("_x_")
    if len(parts) != 2:
        raise ValueError(
            f"axis_layout invalide : « {axis_layout} ». "
            "Format attendu : « dimA_x_dimB » (ex : brands_x_modalites)."
        )
    cat_dim = _normalize_dim(parts[0])
    ser_dim = _normalize_dim(parts[1])
    if cat_dim == ser_dim:
        raise ValueError(
            f"axis_layout « {axis_layout} » : catégories et séries distinctes requises."
        )
    return ({"brand", "modality", "segment"} - {cat_dim, ser_dim}).pop()


def list_unique_values_for_dimension(data: SlideData, dimension: str) -> list[str]:
    """
    Valeurs uniques (ordre de première apparition, hors « Total »).

    - ``segment`` : noms de séries (en-têtes) — filtrage ``filter_segment``.
    - ``brand`` / ``modality`` : colonnes A / B du tableur source.
    """
    d = (dimension or "").strip().lower()
    if d == "brand":
        if not data.brands:
            return []
        return list(
            dict.fromkeys(
                b for b in data.brands
                if b and str(b).strip() and str(b).strip().lower() != "total"
            )
        )
    if d == "modality":
        if not data.subcategories:
            return []
        return list(
            dict.fromkeys(
                m for m in data.subcategories
                if m and str(m).strip() and str(m).strip().lower() != "total"
            )
        )
    if d == "segment":
        return list(dict.fromkeys(s for s in data.series_names if s and str(s).strip()))
    raise ValueError(
        f"Dimension inconnue : « {dimension} ». Attendu : brand, modality, segment."
    )


def _filter_indices(items: list[str], filt: str, dim_label: str) -> list[int]:
    """Stratégie en cascade : match exact → préfixe → contenance."""
    if not filt or filt.strip().lower() in ("", "all"):
        return list(range(len(items)))
    f = filt.strip().lower()
    normalized = [(x or "").strip().lower() for x in items]

    exact = [i for i, x in enumerate(normalized) if x == f]
    if exact:
        return exact
    prefix = [i for i, x in enumerate(normalized) if x.startswith(f)]
    if prefix:
        return prefix
    contains = [i for i, x in enumerate(normalized) if f in x]
    if contains:
        return contains

    raise ValueError(
        f"filter_{dim_label} = « {filt} » ne correspond à aucune valeur. "
        f"Valeurs disponibles : {items}"
    )


def apply_axis_layout(
    data: SlideData,
    *,
    axis_layout: str,
    filter_brand: str = "",
    filter_modalite: str = "",
    filter_segment: str = "",
) -> SlideData:
    """Pivot générique : catégories × séries, avec filtres par dimension."""
    parts = axis_layout.strip().lower().split("_x_")
    if len(parts) != 2:
        raise ValueError(
            f"axis_layout invalide : « {axis_layout} ». "
            "Format attendu : « dimA_x_dimB » (ex : brands_x_modalites)."
        )
    cat_dim = _normalize_dim(parts[0])
    ser_dim = _normalize_dim(parts[1])
    if cat_dim == ser_dim:
        raise ValueError(
            f"axis_layout « {axis_layout} » : catégories et séries "
            "ne peuvent pas être la même dimension."
        )

    other_dim = ({"brand", "modality", "segment"} - {cat_dim, ser_dim}).pop()

    brands_raw = data.brands or [""] * len(data.categories)
    modalities_raw = data.subcategories or [""] * len(data.categories)
    segments_raw = list(data.series_names)

    u_brands: list[str] = list(dict.fromkeys(b for b in brands_raw if b)) or [""]
    u_modalities: list[str] = list(dict.fromkeys(m for m in modalities_raw if m)) or [""]
    u_segments: list[str] = segments_raw

    b_pos = {b: i for i, b in enumerate(u_brands)}
    m_pos = {m: i for i, m in enumerate(u_modalities)}
    cube = [[[0.0] * len(u_segments) for _ in u_modalities] for _ in u_brands]
    for ri, (b, m) in enumerate(zip(brands_raw, modalities_raw)):
        bi = b_pos.get(b if b else "", 0)
        mi = m_pos.get(m if m else "", 0)
        for si in range(len(u_segments)):
            cube[bi][mi][si] = data.values_matrix[si][ri]

    b_idx = _filter_indices(u_brands, filter_brand, "brand")
    m_idx = _filter_indices(u_modalities, filter_modalite, "modalite")
    s_idx = _filter_indices(u_segments, filter_segment, "segment")

    kept = {"brand": b_idx, "modality": m_idx, "segment": s_idx}
    names = {"brand": u_brands, "modality": u_modalities, "segment": u_segments}

    if len(kept[other_dim]) != 1:
        raise ValueError(
            f"Avec axis_layout = « {axis_layout} », la dimension "
            f"« {other_dim} » doit être filtrée à une seule valeur via "
            f"filter_{other_dim}. Actuellement : {len(kept[other_dim])} valeur(s) "
            f"retenue(s) parmi {names[other_dim]}."
        )

    cat_labels = [names[cat_dim][i] for i in kept[cat_dim]]
    ser_labels = [names[ser_dim][i] for i in kept[ser_dim]]

    values_matrix: list[list[float]] = []
    for s_i in kept[ser_dim]:
        row_vals: list[float] = []
        for c_i in kept[cat_dim]:
            coord = {cat_dim: c_i, ser_dim: s_i, other_dim: kept[other_dim][0]}
            row_vals.append(cube[coord["brand"]][coord["modality"]][coord["segment"]])
        values_matrix.append(row_vals)

    return SlideData(
        series_names=ser_labels,
        categories=cat_labels,
        values_matrix=values_matrix,
        brands=None,
        subcategories=None,
    )


# ═════════════════════════════════════════════════════════════════════════════
# 5. build_chart_data — prépare CategoryChartData à partir de SlideData
# ═════════════════════════════════════════════════════════════════════════════

def _normalize_matrix_per_category_pct(matrix: list[list[float]]) -> list[list[float]]:
    """Normalise chaque colonne catégorie à 100 % (pour stacked_100)."""
    if not matrix or not matrix[0]:
        return matrix
    n_ser = len(matrix)
    n_cat = len(matrix[0])
    out: list[list[float]] = [[0.0] * n_cat for _ in range(n_ser)]
    for j in range(n_cat):
        total = sum(matrix[i][j] for i in range(n_ser))
        if total <= 0:
            for i in range(n_ser):
                out[i][j] = matrix[i][j]
            continue
        for i in range(n_ser):
            out[i][j] = matrix[i][j] / total * 100.0
    return out


def build_chart_data(
    data: SlideData,
    *,
    normalize_stacked_100: bool = False,
) -> CategoryChartData:
    """Construit un objet :class:`CategoryChartData` prêt à être inséré."""
    matrix = data.values_matrix
    if normalize_stacked_100:
        matrix = _normalize_matrix_per_category_pct(matrix)
    chart_data = CategoryChartData()
    chart_data.categories = data.categories
    for name, vals in zip(data.series_names, matrix):
        chart_data.add_series(name, vals)
    return chart_data
